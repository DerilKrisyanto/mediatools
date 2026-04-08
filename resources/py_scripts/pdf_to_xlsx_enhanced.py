#!/usr/bin/env python3
"""Enhanced PDF → XLSX with better table detection"""

import sys
import json
import traceback

def convert_with_enhanced_detection(pdf_path, xlsx_path):
    try:
        import pdfplumber
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        wb.remove(wb.active)
        
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                ws = wb.create_sheet(f"Page {page_num + 1}")
                
                tables = page.extract_tables()
                
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        
                        # Write table data
                        for row_idx, row_data in enumerate(table):
                            for col_idx, cell_value in enumerate(row_data or []):
                                cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                                cell.value = str(cell_value or '').strip()
                                
                                # Header styling
                                if row_idx == 0:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                else:
                                    # Alternate row colors
                                    if row_idx % 2 == 0:
                                        cell.fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
                                    
                                    # Center numeric values
                                    if cell.value and cell.value.isdigit():
                                        cell.alignment = Alignment(horizontal="center")
                                
                                # Add borders
                                cell.border = Border(
                                    left=Side(style='thin', color='CCCCCC'),
                                    right=Side(style='thin', color='CCCCCC'),
                                    top=Side(style='thin', color='CCCCCC'),
                                    bottom=Side(style='thin', color='CCCCCC')
                                )
                        
                        # Auto-adjust column widths
                        for col in ws.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = min(50, max(12, max_length + 2))
                            ws.column_dimensions[column].width = adjusted_width
        
        if not wb.worksheets:
            ws = wb.create_sheet("Results")
            ws['A1'] = "No tables found in PDF"
        
        wb.save(xlsx_path)
        
        return {
            'success': True,
            'method': 'enhanced-pdfplumber-openpyxl',
            'sheets': len(wb.worksheets)
        }
        
    except ImportError as e:
        return {
            'success': False,
            'error': f'Missing library: {str(e)}'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({'success': False, 'error': 'Usage: pdf_to_xlsx_enhanced.py <input.pdf> <output.xlsx>'}))
        sys.exit(1)
    
    result = convert_with_enhanced_detection(sys.argv[1], sys.argv[2])
    print(json.dumps(result))
    sys.exit(0 if result.get('success') else 1)